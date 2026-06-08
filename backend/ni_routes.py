"""RQIA Registered Services — Northern Ireland (OpenDataNI dataset).

The Regulation and Quality Improvement Authority (RQIA) publishes a
monthly XLSX of every regulated service in NI via OpenDataNI's CKAN
portal. There is no live API — we either fetch the latest file from
CKAN on demand, or accept a manual upload (same pattern as
``scotland_routes``).

Collections:
  • ``ni_care_services``       — one document per service (id derived
    from the trailing ``(021098)`` token in Service Name).
  • ``ni_definition``          — single "which services count" rule.
  • ``ni_import_state``        — last-import metadata.

Endpoints (admin only):
  • GET  /api/ni/definition
  • PUT  /api/ni/definition
  • GET  /api/ni/definition/preview
  • GET  /api/ni/distinct?field=...
  • POST /api/ni/import           — multipart XLSX upload
  • POST /api/ni/import/refresh   — fetch latest from OpenDataNI CKAN
  • GET  /api/ni/import/status

Postcode sector + district are derived using the same regex shape used
by the CQC + Scotland importers, so NI postcodes (``BT…``) live in the
same sector namespace as the rest of the UK.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime, timezone
from typing import Optional

import httpx
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from openpyxl import load_workbook

from ni_definition import (
    NiDefinition,
    DEFAULT_DEFINITION_ID,
    definition_to_mongo_filter,
)
from ni_polygons import generate_ni_sector_polygons

logger = logging.getLogger("creative-mojo-admin.ni")

_POSTCODE_RE = re.compile(r"^\s*([A-Z]{1,2}\d[A-Z\d]?)\s*(\d)([A-Z]{2})\s*$", re.I)

# Service Name in the dataset embeds the service id in parentheses,
# e.g. ``Age NI Shared Lives (021098)``. Pull the id out for our stable
# primary key — strip any non-breaking spaces first.
_SERVICE_ID_RE = re.compile(r"\(([A-Za-z0-9]+)\)\s*$")

# OpenDataNI CKAN dataset for "rqia-regulated-services". Resource list
# is one XLSX that the publisher refreshes monthly.
CKAN_DATASET_ID = "8c94d33e-4c4e-47a9-8709-98af8ffa7462"
CKAN_API_URL = (
    "https://admin.opendatani.gov.uk/api/3/action/package_show"
    f"?id={CKAN_DATASET_ID}"
)
# OpenDataNI's CDN returns 403 to bots without a desktop UA.
_HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
}


def parse_sector(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    m = _POSTCODE_RE.match(str(raw).upper())
    if not m:
        return None
    return f"{m.group(1)} {m.group(2)}"


def parse_district(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    m = _POSTCODE_RE.match(str(raw).upper())
    if not m:
        return None
    return m.group(1)


def _clean(v) -> str:
    """Trim whitespace + non-breaking spaces that pepper the source XLSX."""
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").strip()


def _split_categories(raw: str) -> list[str]:
    """Categories of Care is a free-text column where the publisher uses
    a mix of commas, semicolons and newlines as separators."""
    if not raw:
        return []
    # Normalise separators → comma, then split.
    parts = re.split(r"[\n;,]+", raw)
    return [p.strip() for p in parts if p.strip()]


def _row_to_doc(headers: list[str], row: tuple) -> Optional[dict]:
    """Project one XLSX row → persisted document. Header lookup is
    case-insensitive + space-insensitive so the file format can drift
    a little without breaking imports."""
    def col(name: str):
        key = name.lower().replace(" ", "")
        for idx, h in enumerate(headers):
            if h.lower().replace(" ", "") == key:
                return row[idx] if idx < len(row) else None
        return None

    service_name_raw = _clean(col("Service Name"))
    if not service_name_raw:
        return None
    # Strip the trailing "(id)" — display name is cleaner without it.
    m = _SERVICE_ID_RE.search(service_name_raw)
    if m:
        service_id = m.group(1)
        display_name = service_name_raw[:m.start()].strip()
    else:
        # Fall back to the full name as id (sluggified) so we don't drop rows.
        display_name = service_name_raw
        service_id = re.sub(r"[^A-Za-z0-9]+", "-", service_name_raw).strip("-")[:64]
    if not service_id:
        return None

    pc = _clean(col("Postcode")).upper()
    sector = parse_sector(pc)
    district = parse_district(pc)

    places_raw = col("Max Approved Places")
    try:
        max_places = int(float(places_raw)) if places_raw not in (None, "") else None
    except (TypeError, ValueError):
        max_places = None

    last_inspected_raw = col("Last Inspected")
    if isinstance(last_inspected_raw, datetime):
        last_inspected = last_inspected_raw.date().isoformat()
    else:
        last_inspected = _clean(last_inspected_raw)

    return {
        "serviceId": service_id,
        "name": display_name,
        "fullName": service_name_raw,
        "addressLines": [
            _clean(col("Address Line1")),
            _clean(col("Address Line2")),
        ],
        "town": _clean(col("Town")),
        "postalCode": pc,
        "postcode_sector": sector,
        "postcode_district": district,
        "phone": _clean(col("Phone")),
        "serviceType": _clean(col("Service Type")),
        "maxApprovedPlaces": max_places,
        "categoriesOfCare": _split_categories(_clean(col("Categories of Care"))),
        "conditions": _clean(col("Conditions (As per Certificate)")),
        "manager": _clean(col("Current Manager")),
        "provider": _clean(col("Provider")),
        "lastInspectedDate": last_inspected,
        "country": "Northern Ireland",
    }


def _parse_xlsx_bytes(raw: bytes) -> tuple[list[dict], int]:
    """Read XLSX bytes → (docs, skipped). Picks the first sheet.

    Memory is fine — RQIA dataset is ~1.5k rows / 200KB."""
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # pragma: no cover — surfaced to caller
        raise HTTPException(400, detail=f"Could not open XLSX: {e}")
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, detail="XLSX is empty")
    headers = [_clean(h) for h in header_row]
    if "Service Name" not in headers:
        raise HTTPException(
            400,
            detail=(
                "XLSX doesn't look like the RQIA Registered Services file — "
                "expected a 'Service Name' column."
            ),
        )
    docs: list[dict] = []
    skipped = 0
    for row in rows_iter:
        # Skip completely blank rows openpyxl emits at the end of sheets.
        if row is None or all((v is None or str(v).strip() == "") for v in row):
            continue
        d = _row_to_doc(headers, row)
        if d is None:
            skipped += 1
            continue
        docs.append(d)
    return docs, skipped


async def _fetch_latest_url() -> tuple[str, str]:
    """Resolve the current OpenDataNI XLSX URL via CKAN.
    Returns (url, filename)."""
    async with httpx.AsyncClient(timeout=30.0, headers=_HTTP_HEADERS, follow_redirects=True) as client:
        r = await client.get(CKAN_API_URL)
        r.raise_for_status()
        meta = r.json()
    if not meta.get("success"):
        raise HTTPException(502, detail="OpenDataNI CKAN API returned a non-success response")
    resources = meta.get("result", {}).get("resources", []) or []
    xlsx = next(
        (res for res in resources if (res.get("format") or "").upper() == "XLSX"),
        None,
    )
    if not xlsx or not xlsx.get("url"):
        raise HTTPException(502, detail="No XLSX resource found in the OpenDataNI dataset")
    url = xlsx["url"]
    filename = url.rsplit("/", 1)[-1] or "rqia_registeredservices.xlsx"
    return url, filename


async def _download_latest_xlsx() -> tuple[bytes, str]:
    url, filename = await _fetch_latest_url()
    # ``follow_redirects`` is needed — OpenDataNI's CKAN endpoint 302s to
    # a presigned Cloudflare R2 URL where the actual file lives.
    async with httpx.AsyncClient(timeout=60.0, headers=_HTTP_HEADERS, follow_redirects=True) as client:
        r = await client.get(url)
        r.raise_for_status()
        return r.content, filename


async def _replace_collection(db, docs: list[dict]) -> None:
    """Atomically swap the live collection for one populated from
    ``docs``. Mirrors the Scotland importer."""
    await db.ni_care_services_tmp.drop()
    CHUNK = 5000
    for i in range(0, len(docs), CHUNK):
        await db.ni_care_services_tmp.insert_many(docs[i:i + CHUNK])
    await db.ni_care_services_tmp.create_index("serviceId", unique=True)
    await db.ni_care_services_tmp.create_index("postcode_sector")
    await db.ni_care_services_tmp.create_index("serviceType")
    await db.ni_care_services_tmp.create_index("categoriesOfCare")
    await db.ni_care_services_tmp.create_index("provider")
    await db.ni_care_services_tmp.create_index("town")
    await db.ni_care_services.drop()
    await db.ni_care_services_tmp.rename("ni_care_services")


# ----------------------------------------------------------------- router
def build_ni_router(db, require_role):  # noqa: D401
    router = APIRouter()

    async def _get_def() -> NiDefinition:
        doc = await db.ni_definition.find_one({"_id": DEFAULT_DEFINITION_ID}, {"_id": 0})
        if not doc:
            return NiDefinition()
        return NiDefinition(**doc)

    async def _regen_polygons_safely() -> Optional[dict]:
        """Regenerate BT sector polygons. Logs and swallows any
        upstream errors (postcodes.io down etc.) so an XLSX import
        doesn't fail just because the polygon refresh did."""
        try:
            return await generate_ni_sector_polygons(db)
        except Exception as e:  # pragma: no cover — defensive
            logger.warning("NI polygon regeneration failed: %s", e)
            return None

    async def _recount_ni_franchisees(ni_def: NiDefinition) -> int:
        """Re-derive ``territory_home_count`` for every franchisee with
        BT sectors in their territory, mixing in their non-NI sector
        counts from CQC + Scotland so the headline number stays whole.

        Lazy-imports the CQC + Scotland leaf modules to avoid a router
        load-order dependency (this module gets registered before they
        are guaranteed to be importable in some env configs)."""
        from cqc_definition import (
            CqcDefinition,
            DEFAULT_DEFINITION_ID as CQC_ID,
            definition_to_mongo_filter as cqc_filter,
        )
        from scotland_definition import (
            ScotlandDefinition,
            DEFAULT_DEFINITION_ID as SCOT_ID,
            definition_to_mongo_filter as scot_filter,
        )
        from geo_postcode import is_scottish_postcode

        cqc_doc = await db.cqc_definition.find_one({"_id": CQC_ID}, {"_id": 0})
        cqc_def = CqcDefinition(**cqc_doc) if cqc_doc else CqcDefinition()
        scot_doc = await db.scotland_definition.find_one({"_id": SCOT_ID}, {"_id": 0})
        scot_def = ScotlandDefinition(**scot_doc) if scot_doc else ScotlandDefinition()

        ni_filter = definition_to_mongo_filter(ni_def)
        cqc_f = cqc_filter(cqc_def)
        scot_f = scot_filter(scot_def)

        updated = 0
        cur = db.franchisees.find(
            {"territory_sectors": {"$exists": True, "$ne": []}},
            {"_id": 0, "id": 1, "territory_sectors": 1},
        )
        async for f in cur:
            sectors = f.get("territory_sectors") or []
            ni_secs = [s for s in sectors if s.upper().startswith("BT")]
            if not ni_secs:
                continue
            other_secs = [s for s in sectors if not s.upper().startswith("BT")]
            scot_secs = [s for s in other_secs if is_scottish_postcode(s)]
            cqc_secs = [s for s in other_secs if not is_scottish_postcode(s)]

            total = await db.ni_care_services.count_documents(
                {**ni_filter, "postcode_sector": {"$in": ni_secs}}
            )
            if scot_secs:
                total += await db.scotland_care_services.count_documents(
                    {**scot_f, "postcode_sector": {"$in": scot_secs}}
                )
            if cqc_secs:
                live = await db.cqc_locations_live.count_documents(
                    {**cqc_f, "postcode_sector": {"$in": cqc_secs}}
                )
                if live == 0:
                    live = await db.cqc_locations.count_documents(
                        {**cqc_f, "postcode_sector": {"$in": cqc_secs}}
                    )
                total += live
            await db.franchisees.update_one(
                {"id": f["id"]}, {"$set": {"territory_home_count": total}}
            )
            updated += 1
        return updated

    @router.get("/ni/definition")
    async def get_definition(_user: dict = Depends(require_role("admin"))):
        d = await _get_def()
        return d.model_dump()

    @router.put("/ni/definition")
    async def put_definition(body: NiDefinition, user: dict = Depends(require_role("admin"))):
        doc = body.model_dump()
        doc.update({
            "_id": DEFAULT_DEFINITION_ID,
            "updated_at": datetime.now(timezone.utc),
            "updated_by": user.get("email"),
        })
        await db.ni_definition.update_one(
            {"_id": DEFAULT_DEFINITION_ID}, {"$set": doc}, upsert=True,
        )
        franchisees_updated = await _recount_ni_franchisees(body)
        return {**body.model_dump(), "_recount": {"franchisees_updated": franchisees_updated}}

    @router.get("/ni/definition/preview")
    async def preview_definition(
        include_service_types: Optional[str] = Query(None),
        exclude_service_types: Optional[str] = Query(None),
        include_categories: Optional[str] = Query(None),
        exclude_categories: Optional[str] = Query(None),
        include_providers: Optional[str] = Query(None),
        min_places: Optional[str] = Query(None),
        _user: dict = Depends(require_role("admin")),
    ):
        def split(v):
            return [x.strip() for x in (v or "").split(",") if x.strip()]

        def _opt_int(v):
            try:
                return int(v) if v and str(v).strip() else None
            except (TypeError, ValueError):
                return None

        d = NiDefinition(
            include_service_types=split(include_service_types),
            exclude_service_types=split(exclude_service_types),
            include_categories=split(include_categories),
            exclude_categories=split(exclude_categories),
            include_providers=split(include_providers),
            min_places=_opt_int(min_places),
        )
        f = definition_to_mongo_filter(d)
        count = await db.ni_care_services.count_documents(f)
        by_town = await db.ni_care_services.aggregate([
            {"$match": f},
            {"$group": {"_id": "$town", "n": {"$sum": 1}}},
            {"$sort": {"n": -1}},
            {"$limit": 12},
        ]).to_list(12)
        sample = await db.ni_care_services.find(
            f,
            {
                "_id": 0, "serviceId": 1, "name": 1, "postalCode": 1, "town": 1,
                "serviceType": 1, "maxApprovedPlaces": 1, "provider": 1,
                "lastInspectedDate": 1,
            },
        ).limit(8).to_list(8)
        return {"count": count, "by_town": by_town, "sample": sample}

    @router.get("/ni/distinct")
    async def distinct_values(
        field: str = Query(..., description="serviceType | categoriesOfCare | provider | town"),
        _user: dict = Depends(require_role("admin")),
    ):
        allowed = {"serviceType", "categoriesOfCare", "provider", "town"}
        if field not in allowed:
            raise HTTPException(400, detail="Invalid field")
        # ``categoriesOfCare`` is an array — unwind so the counts reflect
        # per-category occurrences rather than per-document.
        pipeline: list[dict]
        if field == "categoriesOfCare":
            pipeline = [
                {"$unwind": f"${field}"},
                {"$group": {"_id": f"${field}", "n": {"$sum": 1}}},
                {"$sort": {"n": -1}},
                {"$limit": 200},
            ]
        else:
            pipeline = [
                {"$group": {"_id": f"${field}", "n": {"$sum": 1}}},
                {"$sort": {"n": -1}},
                {"$limit": 200},
            ]
        rows = await db.ni_care_services.aggregate(pipeline).to_list(500)
        return {"values": [{"value": r["_id"], "count": r["n"]} for r in rows if r["_id"]]}

    @router.post("/ni/import")
    async def import_xlsx(
        file: UploadFile = File(...),
        user: dict = Depends(require_role("admin")),
    ):
        """Wipe + reload ``ni_care_services`` from an uploaded RQIA XLSX."""
        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(400, detail="Please upload a .xlsx file")
        raw = await file.read()
        docs, skipped = _parse_xlsx_bytes(raw)
        if not docs:
            raise HTTPException(400, detail="No usable rows found — does the file have a 'Service Name' column?")
        await _replace_collection(db, docs)
        meta = {
            "_id": "last_import",
            "filename": file.filename,
            "source": "upload",
            "imported_at": datetime.now(timezone.utc),
            "imported_by": user.get("email"),
            "rows_loaded": len(docs),
            "rows_skipped": skipped,
        }
        await db.ni_import_state.update_one({"_id": "last_import"}, {"$set": meta}, upsert=True)
        polygons = await _regen_polygons_safely()
        return {"ok": True, "rows_loaded": len(docs), "rows_skipped": skipped, "filename": file.filename, "polygons": polygons}

    @router.post("/ni/import/refresh")
    async def refresh_from_opendatani(user: dict = Depends(require_role("admin"))):
        """Pull the current XLSX straight from OpenDataNI's CKAN portal
        (the publisher refreshes the resource monthly) and reload."""
        try:
            raw, filename = await _download_latest_xlsx()
        except httpx.HTTPError as e:
            raise HTTPException(502, detail=f"Could not reach OpenDataNI: {e}")
        docs, skipped = _parse_xlsx_bytes(raw)
        if not docs:
            raise HTTPException(502, detail="OpenDataNI file parsed to zero rows — schema may have changed")
        await _replace_collection(db, docs)
        meta = {
            "_id": "last_import",
            "filename": filename,
            "source": "opendatani",
            "imported_at": datetime.now(timezone.utc),
            "imported_by": user.get("email"),
            "rows_loaded": len(docs),
            "rows_skipped": skipped,
        }
        await db.ni_import_state.update_one({"_id": "last_import"}, {"$set": meta}, upsert=True)
        polygons = await _regen_polygons_safely()
        return {"ok": True, "rows_loaded": len(docs), "rows_skipped": skipped, "filename": filename, "source": "opendatani", "polygons": polygons}

    @router.post("/ni/polygons/regenerate")
    async def regenerate_polygons(_user: dict = Depends(require_role("admin"))):
        """Rebuild the Voronoi-derived BT postcode sector polygons from
        the current ``ni_care_services`` anchor set. Already runs after
        every import — exposed here for manual re-runs after editing
        anchors or fixing postcodes.io blips."""
        return await generate_ni_sector_polygons(db)

    @router.get("/ni/import/status")
    async def import_status(_user: dict = Depends(require_role("admin"))):
        last = await db.ni_import_state.find_one({"_id": "last_import"}, {"_id": 0})
        live_count = await db.ni_care_services.count_documents({})
        return {"live_count": live_count, "last_import": last}

    return router
